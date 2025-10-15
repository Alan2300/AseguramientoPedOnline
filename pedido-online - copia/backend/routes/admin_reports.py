from flask import Blueprint, request, jsonify, session, make_response, current_app
import datetime as dt
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.styles.numbers import FORMAT_CURRENCY_USD_SIMPLE
from backend.services.report_service import report_orders, report_orders_by_status, report_products

admin_reports_bp = Blueprint("admin_reports", __name__)

def _is_admin() -> bool:
    rol = session.get("rol") or session.get("role")
    if isinstance(rol, str):
        return rol.lower() in {"admin", "administrador"}
    try:
        return int(session.get("id_rol") or 0) == 1
    except Exception:
        return False

def _wb_sheet(title: str, columns, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31] if title else "Reporte"

    head_fill  = PatternFill("solid", fgColor="0D6EFD")
    head_font  = Font(color="FFFFFF", bold=True)
    head_align = Alignment(horizontal="center")
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, c in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col, value=c["header"])
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = head_align
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = max(12, len(c["header"]) + 4)

    money_fmt = FORMAT_CURRENCY_USD_SIMPLE.replace("$", "Q ")

    for r_idx, item in enumerate(rows, start=2):
        for c_idx, c in enumerate(columns, start=1):
            val = item.get(c["key"])
            cell = ws.cell(row=r_idx, column=c_idx)

            fmt = c.get("fmt")
            if fmt == "money":
                try:
                    cell.value = float(val or 0)
                except:
                    cell.value = 0.0
                cell.number_format = money_fmt
            elif fmt == "int":
                try:
                    cell.value = int(val or 0)
                except:
                    cell.value = 0
                cell.number_format = "0"
            elif fmt == "date":
                if isinstance(val, (dt.date, dt.datetime)):
                    cell.value = val
                else:
                    v = None
                    if isinstance(val, str):
                        for f in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                            try:
                                v = dt.datetime.strptime(val, f)
                                break
                            except:
                                pass
                    cell.value = v or val
            else:
                cell.value = val

            cell.border = border

    return wb

def _xlsx_response(wb, filename_prefix: str):
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    resp = make_response(bio.read())
    resp.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    resp.headers["Content-Disposition"] = f'attachment; filename="{filename_prefix}_{stamp}.xlsx"'
    return resp

@admin_reports_bp.route("/api/admin/reports/orders.xlsx", methods=["GET", "OPTIONS"])
def rep_orders_xlsx():
    if request.method == "OPTIONS":
        return ("", 204)
    if not _is_admin():
        return jsonify({"mensaje": "No autorizado"}), 403
    try:
        date_from = request.args.get("from") or None
        date_to   = request.args.get("to") or None
        client_q  = request.args.get("client") or None

        rows = report_orders(date_from, date_to, client_q)
        cols = [
            {"key":"id","header":"ID","fmt":"int"},
            {"key":"fecha","header":"Fecha","fmt":"date"},
            {"key":"cliente_nombre","header":"Cliente"},
            {"key":"cliente_email","header":"Correo"},
            {"key":"estado","header":"Estado"},
            {"key":"productos","header":"Productos","fmt":"int"},
            {"key":"total","header":"Total (Q)","fmt":"money"},
        ]
        wb = _wb_sheet("Pedidos", cols, rows)
        return _xlsx_response(wb, "pedidos")
    except Exception:
        current_app.logger.exception("Error reporte pedidos (xlsx)")
        return jsonify({"message":"Error interno del servidor"}), 500

@admin_reports_bp.route("/api/admin/reports/status.xlsx", methods=["GET", "OPTIONS"])
def rep_status_xlsx():
    if request.method == "OPTIONS":
        return ("", 204)
    if not _is_admin():
        return jsonify({"mensaje": "No autorizado"}), 403
    try:
        status    = request.args.get("status") or "Pendiente"
        date_from = request.args.get("from") or None
        date_to   = request.args.get("to") or None

        rows = report_orders_by_status(status, date_from, date_to)
        cols = [
            {"key":"id","header":"ID","fmt":"int"},
            {"key":"fecha","header":"Fecha","fmt":"date"},
            {"key":"cliente_nombre","header":"Cliente"},
            {"key":"cliente_email","header":"Correo"},
            {"key":"estado","header":"Estado"},
            {"key":"productos","header":"Productos","fmt":"int"},
            {"key":"total","header":"Total (Q)","fmt":"money"},
        ]
        wb = _wb_sheet(f"Pedidos_{status}", cols, rows)
        return _xlsx_response(wb, f"pedidos_{status.replace(' ','_').lower()}")
    except Exception:
        current_app.logger.exception("Error reporte pedidos por estado (xlsx)")
        return jsonify({"message":"Error interno del servidor"}), 500

@admin_reports_bp.route("/api/admin/reports/products.xlsx", methods=["GET", "OPTIONS"])
def rep_products_xlsx():
    if request.method == "OPTIONS":
        return ("", 204)
    if not _is_admin():
        return jsonify({"mensaje": "No autorizado"}), 403
    try:
        kind = request.args.get("filter") or "top"
        if kind not in ("top","low"):
            kind = "top"

        rows = report_products(kind)
        cols = [
            {"key":"id_producto","header":"ID Producto","fmt":"int"},
            {"key":"nombre","header":"Producto"},
            {"key":"cantidad_solicitada","header":"Cantidad solicitada","fmt":"int"},
            {"key":"importe_total","header":"Importe total (Q)","fmt":"money"},
            {"key":"lineas","header":"Líneas","fmt":"int"},
        ]
        nombre = "productos_mas_solicitados" if kind=="top" else "productos_menos_solicitados"
        wb = _wb_sheet("Productos", cols, rows)
        return _xlsx_response(wb, nombre)
    except Exception:
        current_app.logger.exception("Error reporte productos (xlsx)")
        return jsonify({"message":"Error interno del servidor"}), 500
